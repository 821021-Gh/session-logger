"""
Session Report Generator
 Generates an Excel report with bar chart from session_log.csv
 and sends it via email.
"""

import smtplib
import os
import sys
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def generate_excel(csv_path, xlsx_path):
    """Generate Excel report with chart from session log CSV."""
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(csv_path):
        print(f"ERROR: CSV file not found: {csv_path}")
        sys.exit(1)

    # ── Read CSV ───────────────────────────────────────────────────────────
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        header = f.readline().strip()
        if not header.startswith("timestamp"):
            print("ERROR: CSV header not found or invalid format.")
            sys.exit(1)
        for line in f:
            line = line.strip()
            if line:
                rows.append(line.split(","))

    if not rows:
        print("WARNING: No data in CSV file.")
        sys.exit(1)

    # ── Determine date range (last 7 days) ────────────────────────────────
    today = datetime.now().date()
    seven_days_ago = today - timedelta(days=7)

    # Parse rows into structured data
    logins = {}   # account -> {date: [times]}
    logouts = {}  # account -> {date: {login_time: logout_time}}

    for row in rows:
        if len(row) < 4:
            continue
        ts_str, event_type, account, duration = row[0], row[1], row[2], row[3]
        try:
            ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
        except:
            continue
        date = ts.date()
        if date < seven_days_ago:
            continue

        if event_type == "LOGIN":
            if account not in logins:
                logins[account] = {}
            if date not in logins[account]:
                logins[account][date] = []
            logins[account][date].append(ts)
        elif event_type == "LOGOUT":
            if account not in logouts:
                logouts[account] = {}
            # Find the most recent LOGIN before this LOGOUT
            login_key = None
            for d in sorted(logins.get(account, {}).keys(), reverse=True):
                if d <= date:
                    for t in logins[account][d]:
                        if t < ts:
                            login_key = t
                            break
                    if login_key:
                        break
            if login_key:
                duration_min = int((ts - login_key).total_seconds() / 60)
            else:
                duration_min = 0
            if account not in logouts:
                logouts[account] = {}
            if date not in logouts[account]:
                logouts[account][date] = []
            logouts[account][date].append(duration_min)

    # Build per-account-per-day usage table (last 7 days)
    all_accounts = sorted(set(list(logins.keys()) + list(logouts.keys())))
    if "System" in all_accounts:
        all_accounts.remove("System")
    if not all_accounts:
        print("WARNING: No user session data found.")
        sys.exit(1)

    dates = [today - timedelta(days=i) for i in range(6, -1, -1)]
    date_labels = [d.strftime("%m/%d") for d in dates]
    weekday_labels = [d.strftime("%a") for d in dates]

    # Build data matrix: [account][day_index] = total minutes
    data_matrix = {}
    for acc in all_accounts:
        data_matrix[acc] = []
        for d in dates:
            total_mins = 0
            # Sum all sessions for this account on this day
            # LOGOUT duration for this account on this day
            if acc in logouts and d in logouts[acc]:
                total_mins += sum(logouts[acc][d])
            # Also check LOGIN entries for display purposes
            data_matrix[acc].append(total_mins)

    # ── Create Excel Workbook ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Session Report"

    # Color palette
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    acc_colors = [
        "2E75B6", "70AD47", "ED7D31", "FFC000",
        "9E479E", "5B9BD5", "839ED0", "9B7FCA"
    ]
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Title row
    report_date = today.strftime("%Y-%m-%d")
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Windows Session Report — Last 7 Days ({report_date})"
    title_cell.font = Font(color="1F4E79", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle row
    ws.merge_cells("A2:J2")
    subtitle = ws["A2"]
    subtitle.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Account: {os.environ.get('USERNAME','Unknown')}"
    subtitle.font = Font(italic=True, size=10, color="888888")
    subtitle.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Blank row
    ws.row_dimensions[3].height = 8

    # Column headers (row 4)
    header_row = 4
    ws.cell(row=header_row, column=1, value="Account")
    ws.cell(row=header_row, column=1).font = header_font
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=1).border = thin_border

    for i, (dl, wl) in enumerate(zip(date_labels, weekday_labels)):
        col = i + 2
        ws.cell(row=header_row, column=col, value=f"{wl}\n{dl}")
        ws.cell(row=header_row, column=col).font = header_font
        ws.cell(row=header_row, column=col).fill = header_fill
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.cell(row=header_row, column=col).border = thin_border

    # Total column
    ws.cell(row=header_row, column=len(dates) + 2, value="Total")
    ws.cell(row=header_row, column=len(dates) + 2).font = header_font
    ws.cell(row=header_row, column=len(dates) + 2).fill = header_fill
    ws.cell(row=header_row, column=len(dates) + 2).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=len(dates) + 2).border = thin_border

    ws.row_dimensions[header_row].height = 30

    # Data rows
    for row_idx, acc in enumerate(all_accounts):
        r = row_idx + header_row + 1
        color = acc_colors[row_idx % len(acc_colors)]

        # Account name
        name_cell = ws.cell(row=r, column=1, value=acc)
        name_cell.font = Font(bold=True, color="FFFFFF")
        name_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        name_cell.border = thin_border

        total = 0
        for day_idx, d in enumerate(dates):
            col = day_idx + 2
            mins = data_matrix[acc][day_idx]
            total += mins
            cell = ws.cell(row=r, column=col)
            if mins > 0:
                cell.value = f"{int(mins // 60)}h {int(mins % 60)}m" if mins >= 60 else f"{int(mins)}m"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(color="AAAAAA")
            cell.fill = PatternFill(start_color="F2F7FB" if row_idx % 2 == 0 else "FFFFFF",
                                    end_color="F2F7FB" if row_idx % 2 == 0 else "FFFFFF",
                                    fill_type="solid")
            cell.border = thin_border

        # Total column
        total_cell = ws.cell(row=r, column=len(dates) + 2)
        if total > 0:
            total_cell.value = f"{int(total // 60)}h {int(total % 60)}m" if total >= 60 else f"{int(total)}m"
            total_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        else:
            total_cell.value = "-"
            total_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_cell.font = Font(bold=True, color="1F4E79")
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.border = thin_border

        ws.row_dimensions[r].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 14
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11
    ws.column_dimensions[get_column_letter(len(dates) + 2)].width = 12

    # ── Add Chart ──────────────────────────────────────────────────────────
    chart_start_row = header_row + len(all_accounts) + 2
    chart_title_row = chart_start_row - 1

    # Summary section for chart data
    chart_data_row = chart_start_row + 1

    ws.cell(row=chart_title_row, column=1, value="Usage Chart (minutes/day)")
    ws.cell(row=chart_title_row, column=1).font = Font(bold=True, size=12, color="1F4E79")

    # Chart data: account | day1 | day2 | ... | day7
    # Write compact data table for chart
    ws.cell(row=chart_data_row, column=1, value="Account")
    for i, dl in enumerate(date_labels):
        ws.cell(row=chart_data_row, column=i + 2, value=dl)

    for row_idx, acc in enumerate(all_accounts):
        r = chart_data_row + 1 + row_idx
        ws.cell(row=r, column=1, value=acc)
        for day_idx, d in enumerate(dates):
            ws.cell(row=r, column=day_idx + 2, value=data_matrix[acc][day_idx])

    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Daily Usage per Account (minutes)"
    chart.style = 10
    chart.y_axis.title = "Minutes"
    chart.x_axis.title = "Date"
    chart.width = 20
    chart.height = 12

    data_ref = Reference(ws,
                         min_col=2, max_col=len(dates) + 1,
                         min_row=chart_data_row,
                         max_row=chart_data_row + len(all_accounts))
    cats_ref = Reference(ws, min_col=2, max_col=len(dates) + 1, min_row=chart_data_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    # Place chart to the right of data table
    chart_col = get_column_letter(len(dates) + 4)
    ws.add_chart(chart, f"{chart_col}{chart_data_row}")

    # ── Footer ─────────────────────────────────────────────────────────────
    footer_row = chart_data_row + len(all_accounts) + 3
    ws.cell(row=footer_row, column=1, value=f"SessionLogger — spunk.chang@gmail.com — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, size=9, color="888888")

    wb.save(xlsx_path)
    print(f"Excel report saved: {xlsx_path}")
    return xlsx_path, all_accounts, data_matrix, date_labels


def send_email(xlsx_path, account_names, data_matrix, date_labels, to_addr):
    """Send email with Excel attachment via Gmail SMTP."""
    import getpass

    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587

    print(f"\nSending email to: {to_addr}")
    print("Note: App Password required (not your regular Gmail password)")
    print("Generate App Password: https://myaccount.google.com/apppasswords")
    print()

    from_addr = to_addr  # same as sender

    try:
        app_password = os.environ.get("GMAIL_APP_PASSWORD")
        if not app_password:
            app_password = getpass.getpass("Gmail App Password: ").strip()
    except Exception:
        app_password = os.environ.get("GMAIL_APP_PASSWORD", "")

    if not app_password:
        print("ERROR: No Gmail App Password provided.")
        print("Generate one at: https://myaccount.google.com/apppasswords")
        sys.exit(1)

    # Build summary text
    total_usage = {}
    for acc in account_names:
        total_usage[acc] = sum(data_matrix[acc])

    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg["Subject"] = f"Windows Session Report — {datetime.now().strftime('%Y-%m-%d')}"

    # HTML body
    html_body = f"""
    <html><body style="font-family: Arial, sans-serif; color: #333;">
    <h2 style="color:#1F4E79;">Windows Session Report</h2>
    <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    <p>Period: Last 7 days</p>
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse:collapse; font-size:13px;">
    <tr style="background:#1F4E79; color:white;">
      <th>Account</th>
      <th>Total Time</th>
      <th>Days Active</th>
    </tr>
    """

    for acc in account_names:
        total_mins = total_usage.get(acc, 0)
        h = int(total_mins // 60)
        m = int(total_mins % 60)
        days_active = sum(1 for d in data_matrix[acc] if d > 0)
        bg = "#F2F7FB" if account_names.index(acc) % 2 == 0 else "#FFFFFF"
        html_body += f"""<tr style="background:{bg};">
          <td><b>{acc}</b></td>
          <td>{h}h {m}m</td>
          <td>{days_active} days</td>
        </tr>"""

    html_body += "</table><p>See the attached Excel file for the detailed chart.</p></body></html>"

    msg.attach(MIMEText(html_body, "html"))

    # Attach Excel
    if os.path.exists(xlsx_path):
        with open(xlsx_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(xlsx_path)}")
        msg.attach(part)

    # Send
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(from_addr, app_password)
        server.sendmail(from_addr, [to_addr], msg.as_string())
        server.quit()
        print(f"Email sent successfully to {to_addr}")
        return True
    except smtplib.SMTPAuthenticationError:
        print("AUTH FAILED: Invalid App Password. Get one at: https://myaccount.google.com/apppasswords")
        return False
    except Exception as e:
        print(f"Email send failed: {e}")
        return False


if __name__ == "__main__":
    csv_path = os.path.join(os.environ["USERPROFILE"], "session_log.csv")
    xlsx_path = os.path.join(os.environ["USERPROFILE"], "session_report.xlsx")
    to_addr = "spunk.chang@gmail.com"

    if len(sys.argv) > 1 and sys.argv[1] == "--email-only":
        # Called by scheduler, generate + send
        xlsx_path_out, acc_names, data_mat, dl = generate_excel(csv_path, xlsx_path)
        send_email(xlsx_path_out, acc_names, data_mat, dl, to_addr)
    elif len(sys.argv) > 1 and sys.argv[1] == "--generate-only":
        generate_excel(csv_path, xlsx_path)
    else:
        print("Usage:")
        print("  --generate-only  Generate Excel without sending")
        print("  --email-only     Generate Excel and send email")